"""
CSV processor for bike references.
Reads and processes bike references from Conway CSV file.
"""

import csv
import logging
from typing import List, Set, Dict, Any, Optional
from pathlib import Path
from config.settings import settings
from utils.dropbox_handler import get_conway_csv_file

try:
    import openpyxl
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

logger = logging.getLogger(__name__)

class CSVProcessor:
    """
    Processes Conway bike references CSV file.
    Extracts bike references and provides filtering capabilities.
    """
    
    def __init__(self, csv_file_path: str = None):
        """
        Initialize CSV processor.
        
        Args:
            csv_file_path: Path to CSV file. If not provided, will download from Dropbox.
        """
        self.csv_file_path = csv_file_path
        self.bike_references: Set[str] = set()
        self.csv_data: List[Dict[str, str]] = []
        self.temp_file_path: Optional[str] = None
        
        # Load bike references on initialization
        self._load_bike_references()
    
    def _load_bike_references(self) -> None:
        """
        Load bike references from CSV file.
        Downloads from Dropbox if no local file provided.
        Extracts references from the 'Artikelnummer' column.
        """
        try:
            # If no specific CSV file path provided, download from Dropbox
            if not self.csv_file_path:
                logger.info("Downloading Conway CSV file from Dropbox")
                self.temp_file_path = get_conway_csv_file()
                
                if not self.temp_file_path:
                    raise FileNotFoundError("Failed to download Conway CSV file from Dropbox")
                
                self.csv_file_path = self.temp_file_path
            
            csv_path = Path(self.csv_file_path)
            
            if not csv_path.exists():
                raise FileNotFoundError(f"CSV file not found: {self.csv_file_path}")
            
            logger.info(f"Loading bike references from: {self.csv_file_path}")
            
            # Check if file is Excel or CSV
            file_extension = csv_path.suffix.lower()
            
            if file_extension in ['.xlsx', '.xls'] and EXCEL_SUPPORT:
                self._load_from_excel(csv_path)
            else:
                self._load_from_csv(csv_path)
            
            logger.info(f"Loaded {len(self.bike_references)} bike references from CSV")
            
            # Log first few references for debugging (in test mode)
            if settings.TEST_MODE and self.bike_references:
                sample_refs = list(self.bike_references)[:5]
                logger.debug(f"Sample bike references: {sample_refs}")
                
        except Exception as e:
            logger.error(f"Failed to load bike references from file: {e}")
            raise
    
    def _load_from_excel(self, file_path: Path) -> None:
        """
        Load bike references from Excel file.
        
        Args:
            file_path: Path to the Excel file
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            worksheet = workbook.active
            
            # Find the header row and Artikelnummer column
            artikelnummer_col = None
            header_row = 1
            
            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=header_row, column=col).value
                if cell_value and str(cell_value).lower() in ['artikelnummer', 'article number', 'sku', 'product code']:
                    artikelnummer_col = col
                    break
            
            if artikelnummer_col is None:
                # If no specific column found, try first column
                artikelnummer_col = 1
                logger.warning("Artikelnummer column not found, using first column")
            
            # Process data rows
            for row_num in range(header_row + 1, worksheet.max_row + 1):
                try:
                    cell_value = worksheet.cell(row=row_num, column=artikelnummer_col).value
                    
                    if cell_value and str(cell_value).strip():
                        # Clean and normalize the reference
                        clean_reference = str(cell_value).strip()
                        
                        # Add both with and without leading zero for comparison flexibility
                        self.bike_references.add(clean_reference)  # Original format (with leading zero)
                        self.bike_references.add(clean_reference.lstrip('0'))  # Without leading zeros
                        
                        # Store row data for potential future use
                        row_data = {}
                        for col in range(1, worksheet.max_column + 1):
                            header_cell = worksheet.cell(row=header_row, column=col).value
                            data_cell = worksheet.cell(row=row_num, column=col).value
                            if header_cell:
                                row_data[str(header_cell)] = str(data_cell) if data_cell is not None else ''
                        self.csv_data.append(row_data)
                    
                except Exception as e:
                    logger.warning(f"Error processing Excel row {row_num}: {e}")
                    continue
            
        except Exception as e:
            logger.error(f"Failed to load Excel file: {e}")
            raise
    
    def _load_from_csv(self, file_path: Path) -> None:
        """
        Load bike references from CSV file.
        
        Args:
            file_path: Path to the CSV file
        """
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                # Try to detect delimiter (comma or semicolon)
                sample = file.read(1024)
                file.seek(0)
                
                # Determine delimiter based on frequency
                comma_count = sample.count(',')
                semicolon_count = sample.count(';')
                delimiter = ';' if semicolon_count > comma_count else ','
                
                reader = csv.DictReader(file, delimiter=delimiter)
                
                for row_num, row in enumerate(reader, start=2):  # Start at 2 since row 1 is header
                    try:
                        # Extract bike reference from Artikelnummer column
                        reference = row.get('Artikelnummer') or row.get('artikelnummer')
                        
                        if reference and reference.strip():
                            # Clean and normalize the reference
                            clean_reference = reference.strip()
                            
                            # Add both with and without leading zero for comparison flexibility
                            self.bike_references.add(clean_reference)  # Original format (with leading zero)
                            self.bike_references.add(clean_reference.lstrip('0'))  # Without leading zeros
                            
                            # Store full row data for potential future use
                            self.csv_data.append(row)
                        
                    except Exception as e:
                        logger.warning(f"Error processing CSV row {row_num}: {e}")
                        continue
                        
        except UnicodeDecodeError:
            # Try different encodings
            for encoding in ['latin-1', 'cp1252', 'iso-8859-1']:
                try:
                    with open(file_path, 'r', encoding=encoding) as file:
                        sample = file.read(1024)
                        file.seek(0)
                        
                        comma_count = sample.count(',')
                        semicolon_count = sample.count(';')
                        delimiter = ';' if semicolon_count > comma_count else ','
                        
                        reader = csv.DictReader(file, delimiter=delimiter)
                        
                        for row_num, row in enumerate(reader, start=2):
                            try:
                                reference = row.get('Artikelnummer') or row.get('artikelnummer')
                                
                                if reference and reference.strip():
                                    clean_reference = reference.strip()
                                    self.bike_references.add(clean_reference)
                                    self.bike_references.add(clean_reference.lstrip('0'))
                                    self.csv_data.append(row)
                                
                            except Exception as e:
                                logger.warning(f"Error processing CSV row {row_num}: {e}")
                                continue
                    
                    logger.info(f"Successfully loaded CSV with {encoding} encoding")
                    break
                    
                except UnicodeDecodeError:
                    continue
            else:
                raise UnicodeDecodeError("Could not decode file with any supported encoding")
    
    def get_bike_references(self) -> Set[str]:
        """
        Get all loaded bike references.
        
        Returns:
            Set of bike reference strings (normalized to uppercase)
        """
        return self.bike_references.copy()
    
    def contains_bike_reference(self, text: str) -> bool:
        """
        Check if text contains any bike reference.
        
        Args:
            text: Text to search for bike references
            
        Returns:
            True if any bike reference is found in the text
        """
        if not text:
            return False
        
        # Normalize text for comparison (remove extra spaces, but keep original case for numbers)
        normalized_text = ' '.join(text.split())
        
        # Check if any bike reference is present in the text
        for reference in self.bike_references:
            if reference in normalized_text:
                return True
        
        return False
    
    def find_matching_references(self, text: str) -> List[str]:
        """
        Find all bike references present in the given text.
        
        Args:
            text: Text to search for bike references
            
        Returns:
            List of matching bike references found in the text
        """
        if not text:
            return []
        
        # Normalize text for comparison (remove extra spaces, but keep original case for numbers)
        normalized_text = ' '.join(text.split())
        
        # Find all matching references (avoid duplicates by using set)
        matches = set()
        for reference in self.bike_references:
            if reference in normalized_text:
                matches.add(reference)
        
        return list(matches)
    
    def filter_orders_by_references(self, orders: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Filter orders to only include those containing bike references.
        
        Args:
            orders: List of order dictionaries from Holded API
            
        Returns:
            List of orders that contain bike references
        """
        filtered_orders = []
        
        for order in orders:
            try:
                # Check various fields where bike references might appear
                search_fields = [
                    order.get('desc', ''),
                    order.get('notes', ''),
                    order.get('custom', ''),
                ]
                
                # Also check products/line items if present
                if 'products' in order:
                    for product in order.get('products', []):
                        search_fields.extend([
                            product.get('name', ''),
                            product.get('desc', ''),
                            product.get('code', ''),
                            product.get('sku', ''),
                        ])
                elif 'items' in order:
                    for item in order.get('items', []):
                        search_fields.extend([
                            item.get('name', ''),
                            item.get('desc', ''),
                            item.get('code', ''),
                            item.get('sku', ''),
                        ])
                
                # Combine all searchable text
                combined_text = ' '.join(str(field) for field in search_fields if field)
                
                # Check if any bike reference is present
                if self.contains_bike_reference(combined_text):
                    # Add matching references to order data for logging
                    order['matching_references'] = self.find_matching_references(combined_text)
                    filtered_orders.append(order)
                    
            except Exception as e:
                logger.warning(f"Error processing order {order.get('id', 'unknown')}: {e}")
                continue
        
        logger.info(f"Filtered {len(filtered_orders)} orders containing bike references from {len(orders)} total orders")
        
        return filtered_orders
    
    def get_csv_stats(self) -> Dict[str, Any]:
        """
        Get statistics about the loaded CSV data.
        
        Returns:
            Dictionary with CSV statistics
        """
        return {
            'file_path': self.csv_file_path,
            'total_references': len(self.bike_references),
            'total_rows': len(self.csv_data),
            'file_exists': Path(self.csv_file_path).exists() if self.csv_file_path else False,
            'from_dropbox': self.temp_file_path is not None,
        }
    
    def cleanup(self):
        """
        Clean up temporary files if they were downloaded from Dropbox.
        """
        if self.temp_file_path:
            try:
                from utils.dropbox_handler import DropboxHandler
                handler = DropboxHandler()
                handler.cleanup_temp_file(self.temp_file_path)
                self.temp_file_path = None
            except Exception as e:
                logger.warning(f"Failed to cleanup temporary file: {e}")
    
    def __del__(self):
        """Cleanup temporary files when object is destroyed."""
        self.cleanup() 